"""Excel writing helpers for hedging outputs."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_RUNS_DIR = PROJECT_ROOT / "Data" / "runs"
def write_cost_neutral_hedge_results(
    *,
    result: dict,
    instrument_names: list[str],
    loads,
    spot_prices,
    hour_index=None,
    hedge_profile=None,
    forward_prices=None,
    output_dir: str | Path | None = None,
    filename: str | None = None,
) -> Path:
    """Write optimization outputs to a multi-sheet Excel file in Data/runs."""
    out_dir = Path(output_dir) if output_dir is not None else DEFAULT_RUNS_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    # Prepare output directory
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"cost_neutral_hedge_{timestamp}.xlsx"
    filepath = out_dir / filename

    # Validate input
    if hedge_profile is None:
        raise ValueError("hedge_profile must be provided.")
    residual_load = loads - hedge_profile

    # Calculate residual cost: sum(residual_mwh * spot_prices)
    residual_spot_cost_eur = float((pd.Series(residual_load) * pd.Series(spot_prices)).sum())
    # --- Add model flags and parameters at the top of the summary sheet ---
    import inspect
    # Try to import model flags/params from caller if available
    model_flags = {}
    try:
        # Get the caller's global variables (should be cost_neutral_hedge.py)
        frame = inspect.currentframe()
        while frame:
            if frame.f_globals.get("ENFORCE_COVERAGE") is not None:
                break
            frame = frame.f_back
        if frame:
            for key in [
                "ENFORCE_COVERAGE", "HEDGE_RATIO_BOUNDS", "MIN_HEDGE_RATIO",
                "EPSILON", "HEDGE_RATIO_LB", "HEDGE_RATIO_UB", "H_MIN",
                "SUBSET_FILTER", "SAVE_FILTERED_SUBSET",
                "SOLVE_MODE", "CHUNK_OVERLAP_HOURS", "POLISH_GLOBAL_AFTER_CHUNKS",
                "OPTIMIZER_BACKEND",
            ]:
                if key in frame.f_globals:
                    model_flags[key] = frame.f_globals[key]
    except Exception:
        pass

    # Extract model flags/parameters from caller (if available)
    flags_rows = [(k, str(v)) for k, v in model_flags.items()]
    flags_df = pd.DataFrame(flags_rows, columns=["parameter", "value"])

    # Main summary metrics as before
    total_residual_mwh = float(pd.Series(residual_load).sum())
    # Compose summary metrics, inserting constraint values right after 'success'
    summary_metrics = [("success", bool(result["success"]))]
    constraint_values = result.get("constraint_values", [])
    summary_metrics += [(name, val) for name, val in constraint_values]
    summary_metrics += [
        ("message", str(result["message"])),
        ("total_load_mwh", float(pd.Series(loads).sum())),
        ("total_hedged_mwh", float(pd.Series(hedge_profile).sum())),
        ("total_residual_mwh", total_residual_mwh),
        ("hedge_ratio", float(pd.Series(hedge_profile).sum() / (pd.Series(loads).sum() + 1e-10))),
        ("all_to_spot_cost_eur", float((pd.Series(loads) * pd.Series(spot_prices)).sum())),
        ("delta_all_to_spot_vs_forward_hedge_cost_eur", float((pd.Series(loads) * pd.Series(spot_prices)).sum() - result["cost"])),
        ("objective_cost_eur", float(result["cost"])),
        ("residual_spot_cost_eur", residual_spot_cost_eur),
        ("total_costs_eur", float(result["cost"]) + float(residual_spot_cost_eur)),
    ]
    summary_df = pd.DataFrame(summary_metrics, columns=["metric", "value"])

    # Format model flags/params as DataFrame for Excel output
    def try_float_round(x):
        # Keep booleans as-is (e.g., for 'success')
        if isinstance(x, bool):
            return x
        try:
            return round(float(x), 3)
        except Exception:
            return x
    summary_df["value"] = summary_df.apply(
        lambda row: row["value"] if row["metric"] == "success" else try_float_round(row["value"]), axis=1
    )

    # Calculate hedge volumes per instrument (if coverage matrix is available)
    coverage = result.get('coverage', None)
    if coverage is not None:
        import numpy as np
        hedge_volumes = [result["a"][j] * np.sum(coverage[:, j] * loads) for j in range(coverage.shape[1])]
        hours_covered = np.sum(coverage, axis=0)
        hedge_volumes_mw = np.divide(
            np.asarray(hedge_volumes, dtype=float),
            np.asarray(hours_covered, dtype=float),
            out=np.zeros_like(np.asarray(hedge_volumes, dtype=float)),
            where=np.asarray(hours_covered, dtype=float) > 0,
        )
    else:
        hedge_volumes = [None] * len(result["a"])
        hedge_volumes_mw = [None] * len(result["a"])
        hours_covered = [None] * len(result["a"])
    hedge_ratios_df = pd.DataFrame(
        {
            "instrument": instrument_names,
            "hedge_ratio": result["a"],
            "hedge_volume_mwh": hedge_volumes,
            "hedge_volume_mw": hedge_volumes_mw,
            "forward_price_eur_mwh": list(forward_prices) if forward_prices is not None else [None]*len(instrument_names),
            "covered_hours": hours_covered,
        }
    )

    # Prepare hourly profile DataFrame
    hourly_df = pd.DataFrame(
        {
            "load_mwh": loads,
            "spot_price_eur_per_mwh": spot_prices,
            "hedged_mwh": hedge_profile,
            "residual_mwh": residual_load,
        }
    )
    if hour_index is not None:
        ts = pd.to_datetime(hour_index)
        # Excel does not support timezone-aware datetimes.
        if isinstance(ts, pd.DatetimeIndex) and ts.tz is not None:
            # Convert from internal timezone (often UTC) to business-local clock first.
            ts = ts.tz_convert("Europe/Berlin").tz_localize(None)
        hourly_df.insert(0, "timestamp", ts)

    import openpyxl
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Write model flags/params at the top, then a blank row, then summary metrics
        startrow = 0
        if not flags_df.empty:
            # Write model flags/parameters
            flags_df.to_excel(writer, index=False, sheet_name="summary", startrow=startrow)
            startrow += len(flags_df) + 2  # +2 for header and blank row
        # Write summary metrics
        summary_df.to_excel(writer, index=False, sheet_name="summary", startrow=startrow)

        # Write hedge ratios and format relevant columns
        hedge_ratios_df.to_excel(writer, index=False, sheet_name="hedge_ratios")
        ws_ratios = writer.sheets["hedge_ratios"]
        ratios_col_map = {cell.value: cell.column for cell in next(ws_ratios.iter_rows(min_row=1, max_row=1))}
        for col_name in ['hedge_ratio', 'hedge_volume_mwh', 'hedge_volume_mw']:
            col_idx = ratios_col_map.get(col_name, None)
            if col_idx is not None:
                for row in ws_ratios.iter_rows(min_row=2, min_col=1, max_col=ws_ratios.max_column, max_row=ws_ratios.max_row):
                    cell = row[col_idx-1]
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.000'

        # Write hourly profile and format relevant columns
        hourly_df.to_excel(writer, index=False, sheet_name="hourly_profile")
        workbook = writer.book
        ws = writer.sheets["hourly_profile"]
        col_map = {cell.value: cell.column for cell in next(ws.iter_rows(min_row=1, max_row=1))}
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column, max_row=ws.max_row):
            if 'load_mwh' in col_map:
                cell = row[col_map['load_mwh']-1]
                cell.number_format = '#,##0.000'
            if 'spot_price_eur_per_mwh' in col_map:
                cell = row[col_map['spot_price_eur_per_mwh']-1]
                cell.number_format = '#,##0.000'
            if 'hedged_mwh' in col_map:
                cell = row[col_map['hedged_mwh']-1]
                cell.number_format = '#,##0.000'
            if 'residual_mwh' in col_map:
                cell = row[col_map['residual_mwh']-1]
                cell.number_format = '#,##0.000'

        # Format summary sheet: apply German-style number formatting to all numeric values in 'value' column
        ws_summary = writer.sheets["summary"]
        summary_col_map = {cell.value: cell.column for cell in next(ws_summary.iter_rows(min_row=1, max_row=1))}
        value_col = summary_col_map.get('value', None)
        if value_col is not None:
            for row in ws_summary.iter_rows(min_row=2, min_col=1, max_col=ws_summary.max_column, max_row=ws_summary.max_row):
                cell = row[value_col-1]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.000'

    return filepath


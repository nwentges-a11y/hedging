def test_excel_summary_number_format(tmp_path):
    import openpyxl
    # Minimal mock data
    result = {
        "success": True,
        "message": "Test run",
        "cost": 100.0,
        "a": np.array([0.5, 0.5]),
        "coverage": np.array([[1, 0], [0, 1]]),
        "constraint_values": [("Constraint 0 (equality)", 0.0)],
    }
    instrument_names = ["inst1", "inst2"]
    loads = np.array([10.0, 20.0])
    spot_prices = np.array([50.0, 60.0])
    hour_index = pd.date_range("2026-01-01", periods=2, freq="h")
    hedge_profile = np.array([5.0, 10.0])
    forward_prices = np.array([55.0, 65.0])

    # Write Excel file
    output_path = write_cost_neutral_hedge_results(
        result=result,
        instrument_names=instrument_names,
        loads=loads,
        spot_prices=spot_prices,
        hour_index=hour_index,
        hedge_profile=hedge_profile,
        forward_prices=forward_prices,
        output_dir=tmp_path,
        filename="test_output.xlsx"
    )

    # Open the Excel file and check number format in summary sheet
    wb = openpyxl.load_workbook(output_path)
    ws = wb["summary"]
    value_col = None
    for cell in ws[1]:
        if cell.value == "value":
            value_col = cell.column
            break
    assert value_col is not None, "No 'value' column found in summary sheet"

    # Check number format for all numeric cells in 'value' column
    for row in ws.iter_rows(min_row=2, min_col=value_col, max_col=value_col):
        cell = row[0]
        if isinstance(cell.value, (int, float)):
            assert cell.number_format in ['#,##0.000', '#.##0,000'], f"Unexpected number format: {cell.number_format}"
import pytest
import pandas as pd
import numpy as np
from pathlib import Path
from utils.write_excel import write_cost_neutral_hedge_results
import openpyxl

def test_write_cost_neutral_hedge_results(tmp_path):
    # Minimal mock data
    result = {
        "success": True,
        "message": "Test run",
        "cost": 100.0,
        "a": np.array([0.5, 0.5]),
        "coverage": np.array([[1, 0], [0, 1]]),
        "constraint_values": [("Constraint 0 (equality)", 0.0)],
    }
    instrument_names = ["inst1", "inst2"]
    loads = np.array([10.0, 20.0])
    spot_prices = np.array([50.0, 60.0])
    hour_index = pd.date_range("2026-01-01", periods=2, freq="h")
    hedge_profile = np.array([5.0, 10.0])
    forward_prices = np.array([55.0, 65.0])

    # Call the function
    output_path = write_cost_neutral_hedge_results(
        result=result,
        instrument_names=instrument_names,
        loads=loads,
        spot_prices=spot_prices,
        hour_index=hour_index,
        hedge_profile=hedge_profile,
        forward_prices=forward_prices,
        output_dir=tmp_path,
        filename="test_output.xlsx"
    )

    # Check file was created
    assert output_path is not None
    assert Path(output_path).exists()

    # Check Excel content
    wb = openpyxl.load_workbook(output_path)
    assert "summary" in wb.sheetnames
    assert "hedge_ratios" in wb.sheetnames
    assert "hourly_profile" in wb.sheetnames

    # Check summary sheet has expected metrics
    ws = wb["summary"]
    metrics = [cell.value for cell in ws["A"]][1:]  # skip header
    assert "success" in metrics
    assert "objective_cost_eur" in metrics

    # Check formatting (number format for value column)
    value_col = None
    for idx, cell in enumerate(ws[1]):
        if cell.value == "value":
            value_col = idx
    assert value_col is not None
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column, max_row=ws.max_row):
        cell = row[value_col]
        if isinstance(cell.value, (int, float)):
            assert cell.number_format in ("#.##0,000", "#,##0.000")

    # Clean up
    wb.close()
    Path(output_path).unlink()
